"""
Enhanced Exporter for VeloVerify
Supports multiple export formats including Excel, PDF summaries, JSON, and CSV.
"""

import pandas as pd
import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
import logging

# Excel formatting
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# PDF support
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from config import get_config

class ExportFormat:
    """Export format constants."""
    EXCEL = 'excel'
    CSV = 'csv'
    JSON = 'json'
    PDF_SUMMARY = 'pdf_summary'

class EnhancedExcelExporter:
    """Enhanced Excel exporter with professional formatting and charts."""
    
    def __init__(self, config=None):
        self.config = config or get_config()
        self.logger = logging.getLogger(__name__)
        
        # Define professional styles
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, NamedStyle]:
        """Create named styles for consistent formatting."""
        styles = {}
        
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles['header'] = header_style
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(name='Calibri', size=10)
        data_style.alignment = Alignment(horizontal='left', vertical='center')
        data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles['data'] = data_style
        
        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(name='Calibri', size=16, bold=True, color='366092')
        title_style.alignment = Alignment(horizontal='center')
        styles['title'] = title_style
        
        # Summary style
        summary_style = NamedStyle(name="summary")
        summary_style.font = Font(name='Calibri', size=11, bold=True)
        summary_style.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        styles['summary'] = summary_style
        
        return styles
    
    def create_enhanced_excel(self, data_sheets: Dict[str, pd.DataFrame], output_path: str) -> str:
        """Create enhanced Excel file with charts and advanced formatting."""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add styles to workbook
        for style in self.styles.values():
            if style.name not in wb.named_styles:
                wb.add_named_style(style)
        
        # Create summary dashboard first
        self._create_summary_dashboard(wb, data_sheets)
        
        # Create data sheets
        self._create_data_sheets(wb, data_sheets)
        
        # Save workbook
        wb.save(output_path)
        self.logger.info(f"Enhanced Excel file created: {output_path}")
        
        return output_path
    
    def _create_summary_dashboard(self, wb: Workbook, data_sheets: Dict[str, pd.DataFrame]):
        """Create an executive summary dashboard."""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "VeloVerify Processing Summary"
        ws['A1'].style = 'title'
        ws.merge_cells('A1:F1')
        
        # Processing date
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Statistics section
        row = 5
        ws[f'A{row}'] = "PROCESSING STATISTICS"
        ws[f'A{row}'].style = 'summary'
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        # Calculate statistics
        weekly_sheets = {k: v for k, v in data_sheets.items() if k.startswith('Week_Ending_')}
        total_weekly_records = sum(len(df) for df in weekly_sheets.values())
        
        stats = [
            ("Total Weekly Records", total_weekly_records),
            ("Weekly Sheets Created", len(weekly_sheets)),
            ("Duplicates Removed", len(data_sheets.get('Duplicate_Poles_Removed', pd.DataFrame()))),
            ("Records with No Pole", len(data_sheets.get('No_Pole_Allocated', pd.DataFrame()))),
            ("Agent Data Mismatches", len(data_sheets.get('Agent_Data_Mismatches', pd.DataFrame())))
        ]
        
        for stat_name, stat_value in stats:
            ws[f'A{row}'] = stat_name
            ws[f'B{row}'] = stat_value
            ws[f'B{row}'].number_format = '#,##0'
            row += 1
        
        # Weekly breakdown section
        row += 2
        ws[f'A{row}'] = "WEEKLY BREAKDOWN"
        ws[f'A{row}'].style = 'summary'
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Headers for weekly breakdown
        ws[f'A{row}'] = "Week Ending"
        ws[f'B{row}'] = "Records"
        ws[f'C{row}'] = "Percentage"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].style = 'header'
        row += 1
        
        # Weekly data
        chart_data = []
        for sheet_name, df in weekly_sheets.items():
            week_date = sheet_name.replace('Week_Ending_', '')
            record_count = len(df)
            percentage = (record_count / total_weekly_records * 100) if total_weekly_records > 0 else 0
            
            ws[f'A{row}'] = week_date
            ws[f'B{row}'] = record_count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            
            chart_data.append((week_date, record_count))
            row += 1
        
        # Create chart if there's data
        if chart_data and len(chart_data) > 1:
            self._create_weekly_chart(ws, chart_data, row + 2)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_weekly_chart(self, ws, chart_data: List[tuple], start_row: int):
        """Create a chart showing weekly distribution."""
        try:
            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Weekly Records Distribution"
            chart.y_axis.title = 'Number of Records'
            chart.x_axis.title = 'Week Ending'
            
            # Data for chart (assuming chart_data is in the worksheet)
            data_range = Reference(ws, min_col=2, min_row=start_row-len(chart_data), 
                                 max_row=start_row-1, max_col=2)
            categories = Reference(ws, min_col=1, min_row=start_row-len(chart_data), 
                                 max_row=start_row-1)
            
            chart.add_data(data_range, titles_from_data=False)
            chart.set_categories(categories)
            
            # Position chart
            ws.add_chart(chart, f"E{start_row}")
            
        except Exception as e:
            self.logger.warning(f"Could not create chart: {e}")
    
    def _create_data_sheets(self, wb: Workbook, data_sheets: Dict[str, pd.DataFrame]):
        """Create formatted data sheets."""
        # Define sheet order
        sheet_order = []
        
        # Weekly sheets first (chronologically)
        weekly_sheets = [name for name in data_sheets.keys() if name.startswith('Week_Ending_')]
        weekly_sheets.sort(key=lambda x: datetime.strptime(x.split('_')[-1], '%Y-%m-%d'))
        sheet_order.extend(weekly_sheets)
        
        # Quality control sheets
        qc_sheets = ['Duplicate_Poles_Removed', 'No_Pole_Allocated', 'Agent_Data_Mismatches']
        for sheet in qc_sheets:
            if sheet in data_sheets:
                sheet_order.append(sheet)
        
        # Other sheets
        for sheet_name in data_sheets.keys():
            if sheet_name not in sheet_order:
                sheet_order.append(sheet_name)
        
        # Create sheets
        for sheet_name in sheet_order:
            df = data_sheets[sheet_name]
            self._create_formatted_sheet(wb, sheet_name, df)
    
    def _create_formatted_sheet(self, wb: Workbook, sheet_name: str, df: pd.DataFrame):
        """Create a formatted worksheet."""
        # Clean sheet name
        clean_name = self._clean_sheet_name(sheet_name)
        ws = wb.create_sheet(title=clean_name)
        
        if df.empty:
            ws['A1'] = 'No data in this category'
            ws['A1'].style = 'data'
            return
        
        # Add title
        ws['A1'] = sheet_name.replace('_', ' ').title()
        ws['A1'].style = 'title'
        ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        
        # Add data starting from row 3
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header row (row 3 since we have title and blank row)
        header_row = 3
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.style = 'header'
        
        # Format data rows
        for row_num in range(header_row + 1, len(df) + header_row + 1):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.style = 'data'
        
        # Create table
        table_ref = f"A{header_row}:{get_column_letter(len(df.columns))}{len(df) + header_row}"
        table = Table(displayName=f"Table_{clean_name.replace(' ', '_')}", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = f'A{header_row + 1}'
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name to comply with Excel requirements."""
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        clean_name = name
        
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        return clean_name[:31]  # Excel limit

class PDFSummaryExporter:
    """Export processing summary as PDF report."""
    
    def __init__(self, config=None):
        self.config = config or get_config()
        self.logger = logging.getLogger(__name__)
        
        if not HAS_REPORTLAB:
            self.logger.warning("ReportLab not available. PDF export will be disabled.")
    
    def create_pdf_summary(self, data_sheets: Dict[str, pd.DataFrame], output_path: str) -> str:
        """Create PDF summary report."""
        if not HAS_REPORTLAB:
            raise ImportError("ReportLab is required for PDF export. Install with: pip install reportlab")
        
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#366092'),
            alignment=1  # Center
        )
        story.append(Paragraph("VeloVerify Processing Summary", title_style))
        story.append(Spacer(1, 12))
        
        # Generated date
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        # Calculate statistics
        weekly_sheets = {k: v for k, v in data_sheets.items() if k.startswith('Week_Ending_')}
        total_weekly_records = sum(len(df) for df in weekly_sheets.values())
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Weekly Records', f"{total_weekly_records:,}"],
            ['Weekly Sheets Created', str(len(weekly_sheets))],
            ['Duplicates Removed', str(len(data_sheets.get('Duplicate_Poles_Removed', pd.DataFrame())))],
            ['Records with No Pole', str(len(data_sheets.get('No_Pole_Allocated', pd.DataFrame())))],
            ['Agent Data Mismatches', str(len(data_sheets.get('Agent_Data_Mismatches', pd.DataFrame())))]
        ]
        
        summary_table = PDFTable(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Weekly Breakdown
        if weekly_sheets:
            story.append(Paragraph("Weekly Breakdown", styles['Heading2']))
            
            weekly_data = [['Week Ending', 'Records', 'Percentage']]
            for sheet_name, df in weekly_sheets.items():
                week_date = sheet_name.replace('Week_Ending_', '')
                record_count = len(df)
                percentage = (record_count / total_weekly_records * 100) if total_weekly_records > 0 else 0
                weekly_data.append([week_date, f"{record_count:,}", f"{percentage:.1f}%"])
            
            weekly_table = PDFTable(weekly_data)
            weekly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(weekly_table)
        
        # Build PDF
        doc.build(story)
        self.logger.info(f"PDF summary created: {output_path}")
        
        return output_path

class MultiFormatExporter:
    """Main exporter class supporting multiple formats."""
    
    def __init__(self, progress_callback=None, config=None):
        self.config = config or get_config()
        self.progress_callback = progress_callback
        self.logger = logging.getLogger(__name__)
        
        # Initialize format-specific exporters
        self.excel_exporter = EnhancedExcelExporter(config)
        self.pdf_exporter = PDFSummaryExporter(config)
    
    def _update_progress(self, message: str, percentage: float):
        """Update progress callback."""
        if self.progress_callback:
            self.progress_callback(message, percentage)
    
    def export_data(self, data_sheets: Dict[str, pd.DataFrame], output_path: str, 
                   formats: List[str] = None) -> Dict[str, str]:
        """Export data in multiple formats."""
        if formats is None:
            formats = [self.config.get('processing.default_output_format', 'excel')]
        
        results = {}
        base_path = Path(output_path).with_suffix('')
        
        for i, format_type in enumerate(formats):
            progress = 20 + (i / len(formats)) * 60
            
            try:
                if format_type == ExportFormat.EXCEL:
                    self._update_progress("Creating Excel file...", progress)
                    excel_path = f"{base_path}.xlsx"
                    results['excel'] = self.excel_exporter.create_enhanced_excel(data_sheets, excel_path)
                
                elif format_type == ExportFormat.CSV:
                    self._update_progress("Creating CSV files...", progress)
                    csv_folder = f"{base_path}_CSV"
                    results['csv'] = self._export_csv(data_sheets, csv_folder)
                
                elif format_type == ExportFormat.JSON:
                    self._update_progress("Creating JSON file...", progress)
                    json_path = f"{base_path}.json"
                    results['json'] = self._export_json(data_sheets, json_path)
                
                elif format_type == ExportFormat.PDF_SUMMARY:
                    self._update_progress("Creating PDF summary...", progress)
                    pdf_path = f"{base_path}_summary.pdf"
                    results['pdf'] = self.pdf_exporter.create_pdf_summary(data_sheets, pdf_path)
                
            except Exception as e:
                self.logger.error(f"Error exporting {format_type}: {e}")
                continue
        
        self._update_progress("Export complete", 100)
        return results
    
    def _export_csv(self, data_sheets: Dict[str, pd.DataFrame], output_folder: str) -> str:
        """Export data as CSV files."""
        os.makedirs(output_folder, exist_ok=True)
        
        for sheet_name, df in data_sheets.items():
            if not df.empty:
                csv_path = os.path.join(output_folder, f"{sheet_name}.csv")
                df.to_csv(csv_path, index=False)
        
        self.logger.info(f"CSV files created in: {output_folder}")
        return output_folder
    
    def _export_json(self, data_sheets: Dict[str, pd.DataFrame], output_path: str) -> str:
        """Export data as JSON file."""
        json_data = {}
        
        for sheet_name, df in data_sheets.items():
            if not df.empty:
                # Convert DataFrame to records format
                json_data[sheet_name] = df.to_dict('records')
            else:
                json_data[sheet_name] = []
        
        # Add metadata
        json_data['_metadata'] = {
            'export_date': datetime.now().isoformat(),
            'total_sheets': len(data_sheets),
            'format_version': '1.0'
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
        
        self.logger.info(f"JSON file created: {output_path}")
        return output_path
    
    def generate_filename(self, base_path: str = None, format_type: str = 'excel') -> str:
        """Generate timestamped filename."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_format = self.config.get('export.filename_format', 'Lawley_Pole_Permissions_Weekly_{timestamp}')
        
        base_filename = filename_format.format(timestamp=timestamp)
        
        extensions = {
            'excel': '.xlsx',
            'csv': '_CSV',
            'json': '.json',
            'pdf_summary': '_summary.pdf'
        }
        
        extension = extensions.get(format_type, '.xlsx')
        filename = f"{base_filename}{extension}"
        
        if base_path:
            return os.path.join(base_path, filename)
        else:
            return filename

def create_multi_format_exporter(progress_callback=None, config=None) -> MultiFormatExporter:
    """Factory function to create multi-format exporter."""
    return MultiFormatExporter(progress_callback, config) 