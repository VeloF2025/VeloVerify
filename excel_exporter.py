"""
VeloVerify Excel Exporter
Module for creating professionally formatted Excel files from processed data.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from datetime import datetime
import logging
from typing import Dict

class ExcelExporter:
    """Excel export handler with professional formatting."""
    
    def __init__(self, progress_callback=None):
        """Initialize the Excel exporter with optional progress callback."""
        self.progress_callback = progress_callback
        self.logger = logging.getLogger(__name__)
        
        # Define styles
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self.data_font = Font(name='Calibri', size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='center')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _update_progress(self, step: str, percentage: int = 0):
        """Update progress if callback is provided."""
        if self.progress_callback:
            self.progress_callback(step, percentage)
        self.logger.info(f"{step} - {percentage}%")
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame, sheet_name: str):
        """Apply professional formatting to a worksheet."""
        if df.empty:
            return
        
        # Set column widths based on content
        column_widths = {}
        for col_idx, column in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # Calculate optimal width
            max_length = len(str(column))
            for value in df[column].astype(str):
                max_length = max(max_length, len(value))
            
            # Set reasonable limits
            width = min(max(max_length + 2, 10), 50)
            column_widths[column_letter] = width
            worksheet.column_dimensions[column_letter].width = width
        
        # Format header row
        header_row = 1
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.border
        
        # Apply table formatting for better readability
        if len(df) > 0:
            table_ref = f"A1:{worksheet.cell(row=len(df) + 1, column=len(df.columns)).coordinate}"
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_ref)
            
            # Apply table style
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name to comply with Excel requirements."""
        # Excel sheet names cannot be longer than 31 characters
        # Cannot contain: \ / ? * [ ]
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        clean_name = name
        
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        # Truncate if too long
        if len(clean_name) > 31:
            clean_name = clean_name[:31]
        
        return clean_name
    
    def _prepare_dataframe_for_export(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare dataframe for Excel export by cleaning data types."""
        export_df = df.copy()
        
        # Remove any temporary columns used for processing
        columns_to_remove = ['parsed_date', 'week_ending']
        for col in columns_to_remove:
            if col in export_df.columns:
                export_df = export_df.drop(columns=[col])
        
        # Convert datetime columns to strings for better Excel compatibility
        for col in export_df.columns:
            if export_df[col].dtype == 'datetime64[ns]':
                export_df[col] = export_df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Handle any remaining object types that might cause issues
        for col in export_df.columns:
            if export_df[col].dtype == 'object':
                export_df[col] = export_df[col].astype(str)
                # Replace 'nan' strings with empty strings
                export_df[col] = export_df[col].replace('nan', '')
        
        return export_df
    
    def create_excel_file(self, data_sheets: Dict[str, pd.DataFrame], output_path: str) -> str:
        """Create the final Excel file with all sheets and formatting."""
        try:
            self._update_progress("Creating Excel file...", 80)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            total_sheets = len(data_sheets)
            processed_sheets = 0
            
            # Define sheet order for better organization
            sheet_order = []
            
            # Add weekly sheets first (in chronological order)
            weekly_sheets = [name for name in data_sheets.keys() if name.startswith('Week_Ending_')]
            weekly_sheets.sort(key=lambda x: datetime.strptime(x.split('_')[-1], '%Y-%m-%d'))
            sheet_order.extend(weekly_sheets)
            
            # Add quality control sheets
            qc_sheets = ['Duplicate_Poles_Removed', 'No_Pole_Allocated', 'Agent_Data_Mismatches', 'Date_Parse_Errors']
            for sheet in qc_sheets:
                if sheet in data_sheets:
                    sheet_order.append(sheet)
            
            # Add processing summary last
            if 'Processing_Summary' in data_sheets:
                sheet_order.append('Processing_Summary')
            
            # Add any remaining sheets
            for sheet_name in data_sheets.keys():
                if sheet_name not in sheet_order:
                    sheet_order.append(sheet_name)
            
            # Create sheets in order
            for sheet_name in sheet_order:
                df = data_sheets[sheet_name]
                
                # Clean sheet name for Excel compatibility
                clean_name = self._clean_sheet_name(sheet_name)
                
                # Prepare dataframe for export
                export_df = self._prepare_dataframe_for_export(df)
                
                # Create worksheet
                ws = wb.create_sheet(title=clean_name)
                
                # Add data to worksheet
                if not export_df.empty:
                    for r in dataframe_to_rows(export_df, index=False, header=True):
                        ws.append(r)
                    
                    # Apply formatting
                    self._format_worksheet(ws, export_df, clean_name)
                else:
                    # Add a note for empty sheets
                    ws['A1'] = 'No data in this category'
                    ws['A1'].font = self.data_font
                
                processed_sheets += 1
                progress = 80 + (processed_sheets / total_sheets * 15)
                self._update_progress(f"Creating sheet: {clean_name}", int(progress))
            
            # Save the file
            self._update_progress("Saving Excel file...", 95)
            wb.save(output_path)
            
            file_size = os.path.getsize(output_path)
            self.logger.info(f"Excel file created successfully: {output_path}")
            self.logger.info(f"File size: {file_size / 1024 / 1024:.2f} MB")
            self.logger.info(f"Total sheets: {len(wb.sheetnames)}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {str(e)}")
            self._update_progress(f"Excel export error: {str(e)}", 0)
            raise
    
    def generate_filename(self, base_path: str = None) -> str:
        """Generate a timestamped filename for the Excel output."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Lawley_Pole_Permissions_Weekly_{timestamp}.xlsx"
        
        if base_path:
            return os.path.join(base_path, filename)
        else:
            return filename
    
    def validate_export(self, file_path: str, expected_sheets: int = None) -> Dict[str, any]:
        """Validate the exported Excel file."""
        try:
            # Load the file to verify it's valid
            with pd.ExcelFile(file_path) as xl:
                sheet_names = xl.sheet_names
                
            validation_results = {
                'file_exists': os.path.exists(file_path),
                'file_size': os.path.getsize(file_path),
                'sheet_count': len(sheet_names),
                'sheet_names': sheet_names,
                'is_valid': True
            }
            
            if expected_sheets and len(sheet_names) != expected_sheets:
                validation_results['is_valid'] = False
                validation_results['error'] = f"Expected {expected_sheets} sheets, got {len(sheet_names)}"
            
            self.logger.info("Excel file validation successful")
            return validation_results
            
        except Exception as e:
            self.logger.error(f"Excel file validation failed: {str(e)}")
            return {
                'file_exists': os.path.exists(file_path) if file_path else False,
                'is_valid': False,
                'error': str(e)
            } 